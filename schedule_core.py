"""
Чтение расписания из Excel (шаблон с группами в шапке) и сохранение в PNG.
Используется локальным ботом (консоль) и при желании — Telegram-ботом.
"""

import os
import re
from collections import defaultdict
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageDraw, ImageFont

# Макет картинки (сетка недели)
IMAGE_WIDTH = 1100
LEFT_PADDING = 24
TOP_PADDING = 20
TITLE_HEIGHT = 48
HEADER_ROW_HEIGHT = 44
TIME_COL_WIDTH = 158
DATA_ROW_HEIGHT = 92  # высота строки пары (несколько строк текста)

# Как на образце: пн–пт и 5 временных окон
WEEKDAYS_RU = ["понедельник", "вторник", "среда", "четверг", "пятница"]
# (начало в минутах от полуночи, конец, подпись в первом столбце)
TIME_SLOTS: List[Tuple[int, int, str]] = [
    (9 * 60 + 0, 10 * 60 + 30, "9:00-10:30"),
    (10 * 60 + 40, 12 * 60 + 10, "10:40-12:10"),
    (12 * 60 + 40, 14 * 60 + 10, "12:40-14:10"),
    (14 * 60 + 20, 15 * 60 + 50, "14:20-15:50"),
    (16 * 60 + 0, 17 * 60 + 30, "16:00-17:30"),
]

# Распознавание дня из ячейки «Понедельник 23.03.2026»
_DAY_WORD_TO_INDEX = {
    "понедельник": 0,
    "вторник": 1,
    "среда": 2,
    "среду": 2,
    "четверг": 3,
    "пятница": 4,
    "пятницу": 4,
}


@dataclass(frozen=True)
class LessonRow:
    day_cell: str
    lesson_no: str
    time_range: str
    subject: str
    teacher: str
    room: str


def _norm_text(value: object) -> str:
    if value is None:
        return ""
    s = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"\s+", " ", s).strip()


def _split_subject_teacher(cell_text: str) -> Tuple[str, str]:
    t = (cell_text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not t:
        return "", ""

    lines = [ln.strip() for ln in t.split("\n") if ln.strip()]
    if len(lines) >= 2:
        return lines[0], " ".join(lines[1:])

    m = re.search(
        r"(?P<subj>.+?)\s+(?P<teacher>[А-ЯA-ZЁ][а-яa-zё\-]+(?:\s+[А-ЯA-ZЁ]\.)+(?:\s+[А-ЯA-ZЁ]\.)?)$",
        lines[0],
    )
    if m:
        return m.group("subj").strip(), m.group("teacher").strip()

    return lines[0], ""


def _is_time_range(text: str) -> bool:
    t = _norm_text(text)
    if not t:
        return False
    return bool(re.search(r"\d{1,2}[:.]\d{2}\s*[-–—]\s*\d{1,2}[:.]\d{2}", t))


def _is_int_like(text: str) -> bool:
    return _norm_text(text).isdigit()


def _build_merged_value_map(ws: Worksheet) -> Dict[Tuple[int, int], str]:
    mapping: Dict[Tuple[int, int], str] = {}
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = mr.min_col, mr.min_row, mr.max_col, mr.max_row
        top_left = ws.cell(row=min_row, column=min_col).value
        val = _norm_text(top_left)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                mapping[(r, c)] = val
    return mapping


def _cell_value(ws: Worksheet, merged_map: Dict[Tuple[int, int], str], row: int, col: int) -> str:
    v = merged_map.get((row, col))
    if v is not None:
        return v
    return _norm_text(ws.cell(row=row, column=col).value)


def _find_group_header_row(ws: Worksheet, scan_rows: int = 8) -> int:
    max_row = min(ws.max_row or 1, scan_rows)
    best_row = 1
    best_score = -1
    for r in range(1, max_row + 1):
        score = 0
        for c in range(1, (ws.max_column or 1) + 1):
            txt = _norm_text(ws.cell(row=r, column=c).value)
            if not txt:
                continue
            if re.search(r"\d", txt) and "-" in txt and len(txt) <= 80:
                score += 2
            if "(" in txt and ")" in txt:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def _parse_groups_from_header(ws: Worksheet, header_row: int) -> Dict[str, Tuple[int, int]]:
    merged_map = _build_merged_value_map(ws)
    groups: Dict[str, Tuple[int, int]] = {}
    max_col = ws.max_column or 1
    c = 1
    while c <= max_col:
        name = _cell_value(ws, merged_map, header_row, c)
        if not name:
            c += 1
            continue
        subj_col = c
        room_col = c + 1
        if room_col > max_col:
            break
        if subj_col <= 3:
            c += 1
            continue
        groups[name] = (subj_col, room_col)
        c += 2
    return groups


def _find_anchor_columns(ws: Worksheet, header_row: int) -> Tuple[int, int]:
    merged_map = _build_merged_value_map(ws)
    scan_start = header_row + 1
    scan_end = min(ws.max_row or scan_start, scan_start + 40)
    lesson_col = 2
    time_col = 3
    best = (-1, lesson_col, time_col)
    for lc in range(1, min(6, (ws.max_column or 6)) + 1):
        for tc in range(1, min(10, (ws.max_column or 10)) + 1):
            if lc == tc:
                continue
            hits = 0
            for r in range(scan_start, scan_end + 1):
                if _is_int_like(_cell_value(ws, merged_map, r, lc)) and _is_time_range(
                    _cell_value(ws, merged_map, r, tc)
                ):
                    hits += 1
            if hits > best[0]:
                best = (hits, lc, tc)
    _, lesson_col, time_col = best
    return lesson_col, time_col


def _find_day_column(
    ws: Worksheet,
    merged_map: Dict[Tuple[int, int], str],
    lesson_col: int,
    time_col: int,
    scan_start_row: int,
    scan_end_row: int,
) -> int:
    max_col = min(6, lesson_col - 1)
    if max_col < 1:
        return 1
    best_col = 1
    best_score = -1
    for dc in range(1, max_col + 1):
        if dc in (lesson_col, time_col):
            continue
        score = 0
        for r in range(scan_start_row, scan_end_row + 1):
            t = _cell_value(ws, merged_map, r, dc)
            if not t:
                continue
            if len(t) >= 8 and any(
                x in t.lower()
                for x in ("понедельник", "вторник", "сред", "четверг", "пятниц", "суббот", "воскрес")
            ):
                score += 3
            if re.search(r"\d{2}\.\d{2}\.\d{4}", t):
                score += 2
        if score > best_score:
            best_score = score
            best_col = dc
    return best_col


def _find_first_data_row(
    ws: Worksheet,
    merged_map: Dict[Tuple[int, int], str],
    header_row: int,
    lesson_col: int,
    time_col: int,
) -> int:
    scan_start = header_row + 1
    scan_end = min(ws.max_row or scan_start, scan_start + 80)
    for r in range(scan_start, scan_end + 1):
        lesson = _cell_value(ws, merged_map, r, lesson_col)
        t = _cell_value(ws, merged_map, r, time_col)
        if _is_time_range(t) and (_is_int_like(lesson) or lesson == ""):
            return r
    return scan_start


def extract_group_schedule(
    xlsx_path: str,
    group_query: str,
    sheet_name: Optional[str] = None,
) -> List[LessonRow]:
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Не найден файл Excel: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    merged_map = _build_merged_value_map(ws)
    header_row = _find_group_header_row(ws)
    groups = _parse_groups_from_header(ws, header_row)

    if not groups:
        raise ValueError(
            "Не удалось найти группы в верхней части листа. "
            "Проверь, что первая строка с названиями групп читается как текст."
        )

    q = group_query.strip()
    q_lower = q.lower()
    matches = [name for name in groups.keys() if q_lower in name.lower()]
    if not matches:
        matches = [name for name in groups.keys() if _norm_text(name).lower() == q_lower]
    if len(matches) != 1:
        preview = "\n".join(f"- {n}" for n in list(groups.keys())[:30])
        extra = "\n... (список обрезан)" if len(groups) > 30 else ""
        if not matches:
            raise ValueError(
                "Группа не найдена.\n"
                f"Ты написал: {q!r}\n\n"
                "Вот примеры названий, которые я вижу в файле:\n"
                f"{preview}{extra}"
            )
        raise ValueError(
            "Слишком много совпадений, уточни название группы.\n"
            f"Подходят: {', '.join(matches)}"
        )

    group_name = matches[0]
    subj_col, room_col = groups[group_name]

    lesson_col, time_col = _find_anchor_columns(ws, header_row)
    data_start = _find_first_data_row(ws, merged_map, header_row, lesson_col, time_col)
    day_col = _find_day_column(
        ws,
        merged_map,
        lesson_col,
        time_col,
        scan_start_row=data_start,
        scan_end_row=min(ws.max_row or data_start, data_start + 200),
    )

    rows_out: List[LessonRow] = []
    max_row = ws.max_row or 1
    current_day = ""

    for r in range(data_start, max_row + 1):
        day_here = _cell_value(ws, merged_map, r, day_col)
        if day_here:
            current_day = day_here

        lesson_no = _cell_value(ws, merged_map, r, lesson_col)
        time_range = _cell_value(ws, merged_map, r, time_col)

        if not _is_time_range(time_range):
            continue
        if not lesson_no:
            lesson_no = ""

        subj_cell = _cell_value(ws, merged_map, r, subj_col)
        room = _cell_value(ws, merged_map, r, room_col)
        if not subj_cell and not room:
            continue

        subject, teacher = _split_subject_teacher(subj_cell)
        rows_out.append(
            LessonRow(
                day_cell=current_day,
                lesson_no=lesson_no,
                time_range=time_range,
                subject=subject,
                teacher=teacher,
                room=room,
            )
        )

    if not rows_out:
        raise ValueError(
            f"Для группы {group_name!r} не нашлось ни одной строки с временем. "
            "Проверь, что лист тот же, что на скриншоте, и что внизу есть строки времени."
        )

    return rows_out


def _windows_font_path(bold: bool) -> Optional[str]:
    windir = os.environ.get("WINDIR", r"C:\Windows")
    fonts_dir = os.path.join(windir, "Fonts")
    name = "arialbd.ttf" if bold else "arial.ttf"
    path = os.path.join(fonts_dir, name)
    return path if os.path.isfile(path) else None


# === ДОБАВЬ В САМЫЙ ВЕРХ (замени функцию load_font полностью) ===

from PIL import ImageFont
import os


def load_font(size: int, bold: bool = False):
    """
    Универсальная загрузка шрифта:
    - Linux (Railway/VPS) → DejaVuSans
    - Windows → Arial
    """

    # 1. Linux (Railway, VPS)
    linux_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    ]

    for path in linux_paths:
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, size=size)
            except:
                pass

    # 2. Windows fallback
    try:
        return ImageFont.truetype("arialbd.ttf" if bold else "arial.ttf", size=size)
    except:
        pass

    # 3. Последний вариант (нежелательно)
    return ImageFont.load_default()


def _parse_weekday_index(day_cell: str) -> Optional[int]:
    """Индекс 0..4 для пн–пт или None (выходной / не распознано)."""
    if not day_cell:
        return None
    t = day_cell.lower().strip()
    # Сначала по словам в тексте (надёжнее, чем \\b с кириллицей)
    words = re.findall(r"[а-яё]+", t)
    for w in words:
        if w in _DAY_WORD_TO_INDEX:
            return _DAY_WORD_TO_INDEX[w]
    # Подстрока: длинные названия раньше
    for word in sorted(_DAY_WORD_TO_INDEX.keys(), key=len, reverse=True):
        if word in t:
            return _DAY_WORD_TO_INDEX[word]
    return None


def _parse_lesson_times_minutes(time_range: str) -> Optional[Tuple[int, int]]:
    s = _norm_text(time_range).replace("–", "-").replace("—", "-")
    m = re.search(
        r"(\d{1,2})[:.](\d{2})\s*-\s*(\d{1,2})[:.](\d{2})",
        s,
    )
    if not m:
        return None
    h1, mi1, h2, mi2 = (int(m.group(i)) for i in range(1, 5))
    start = h1 * 60 + mi1
    end = h2 * 60 + mi2
    if end <= start:
        return None
    return start, end


def _best_slot_index(lesson_start: int, lesson_end: int) -> int:
    best_i = 0
    best_overlap = -1
    for i, (slot_a, slot_b, _) in enumerate(TIME_SLOTS):
        ov = max(0, min(lesson_end, slot_b) - max(lesson_start, slot_a))
        if ov > best_overlap:
            best_overlap = ov
            best_i = i
    if best_overlap <= 0:
        mid = (lesson_start + lesson_end) // 2

        def dist(i: int) -> int:
            sa, sb, _ = TIME_SLOTS[i]
            sm = (sa + sb) // 2
            return abs(mid - sm)

        best_i = min(range(len(TIME_SLOTS)), key=dist)
    return best_i


def _format_cell_line(row: LessonRow) -> str:
    subj = _norm_text(row.subject)
    room = _norm_text(row.room)
    if not subj and _norm_text(row.teacher):
        subj = _norm_text(row.teacher)
    if not subj:
        return ""
    sl = subj.lower()
    if "онлайн" in sl:
        return subj
    if not room:
        return subj
    if room.isdigit():
        return f"{subj} {room} каб."
    rl = room.lower()
    if "зал" in rl or "каб" in rl or "онлайн" in rl:
        return f"{subj} {room}"
    return f"{subj} {room}"


def _lessons_to_week_grid(rows: List[LessonRow]) -> List[List[str]]:
    """5 строк (пары) × 5 столбцов (дни)."""
    cells: Dict[Tuple[int, int], List[str]] = defaultdict(list)
    for row in rows:
        d = _parse_weekday_index(row.day_cell)
        if d is None:
            continue
        span = _parse_lesson_times_minutes(row.time_range)
        if not span:
            continue
        si = _best_slot_index(span[0], span[1])
        line = _format_cell_line(row)
        if line:
            cells[(si, d)].append(line)

    grid: List[List[str]] = []
    for si in range(len(TIME_SLOTS)):
        row_cells = []
        for d in range(len(WEEKDAYS_RU)):
            parts = cells.get((si, d), [])
            # убрать дубликаты, сохранить порядок
            seen = set()
            uniq = []
            for p in parts:
                if p not in seen:
                    seen.add(p)
                    uniq.append(p)
            row_cells.append("\n".join(uniq))
        grid.append(row_cells)
    return grid


def _wrap_lines_to_width(draw: ImageDraw.ImageDraw, text: str, font, max_width: int) -> List[str]:
    if not text.strip():
        return []
    out: List[str] = []
    for paragraph in text.split("\n"):
        words = _norm_text(paragraph).split()
        if not words:
            continue
        current = words[0]
        for w in words[1:]:
            trial = f"{current} {w}"
            bbox = draw.textbbox((0, 0), trial, font=font)
            if bbox[2] - bbox[0] <= max_width:
                current = trial
            else:
                out.append(current)
                current = w
        out.append(current)
    return out


def render_schedule_png(rows: List[LessonRow], title: str) -> str:
    """
    Картинка в формате сетки: строки — пары времени, столбцы — пн–пт (как на образце).
    """
    grid = _lessons_to_week_grid(rows)

    n_slots = len(TIME_SLOTS)
    n_days = len(WEEKDAYS_RU)
    inner_width = IMAGE_WIDTH - 2 * LEFT_PADDING
    data_col_w = (inner_width - TIME_COL_WIDTH) // n_days
    table_w = TIME_COL_WIDTH + data_col_w * n_days
    table_h = HEADER_ROW_HEIGHT + DATA_ROW_HEIGHT * n_slots

    img_h = TOP_PADDING + TITLE_HEIGHT + table_h + TOP_PADDING + 8
    img = Image.new("RGB", (IMAGE_WIDTH, img_h), color="white")
    draw = ImageDraw.Draw(img)

    title_font = load_font(22, bold=False)
    header_font = load_font(14, bold=True)
    time_font = load_font(13, bold=True)
    cell_font = load_font(13, bold=False)

    tb = draw.textbbox((0, 0), title, font=title_font)
    tw = tb[2] - tb[0]
    draw.text(((IMAGE_WIDTH - tw) // 2, TOP_PADDING), title, font=title_font, fill="black")

    table_x0 = LEFT_PADDING + max(0, (inner_width - table_w) // 2)
    table_y0 = TOP_PADDING + TITLE_HEIGHT

    border = (40, 40, 40)

    def cell_rect(col: int, row: int) -> Tuple[int, int, int, int]:
        if row == 0:
            y0 = table_y0
            h = HEADER_ROW_HEIGHT
        else:
            y0 = table_y0 + HEADER_ROW_HEIGHT + (row - 1) * DATA_ROW_HEIGHT
            h = DATA_ROW_HEIGHT
        if col == 0:
            x0 = table_x0
            w = TIME_COL_WIDTH
        else:
            x0 = table_x0 + TIME_COL_WIDTH + (col - 1) * data_col_w
            w = data_col_w
        return x0, y0, x0 + w, y0 + h

    # Заголовки: пустая угловая ячейка + дни
    for c in range(n_days + 1):
        for r in range(n_slots + 1):
            x0, y0, x1, y1 = cell_rect(c, r)
            draw.rectangle([x0, y0, x1, y1], outline=border, width=1)

    # Угол (0,0) — пусто
    # Первая строка: дни
    for d in range(n_days):
        x0, y0, x1, y1 = cell_rect(d + 1, 0)
        name = WEEKDAYS_RU[d]
        bb = draw.textbbox((0, 0), name, font=header_font)
        twd = bb[2] - bb[0]
        th = bb[3] - bb[1]
        tx = x0 + (x1 - x0 - twd) // 2
        ty = y0 + (y1 - y0 - th) // 2
        draw.text((tx, ty), name, font=header_font, fill="black")

    # Столбец времени + тело
    pad = 8
    for si in range(n_slots):
        x0, y0, x1, y1 = cell_rect(0, si + 1)
        label = TIME_SLOTS[si][2]
        lines = _wrap_lines_to_width(draw, label, time_font, TIME_COL_WIDTH - 2 * pad)
        line_h = draw.textbbox((0, 0), "Ay", font=time_font)[3] - draw.textbbox((0, 0), "Ay", font=time_font)[1]
        if not lines:
            lines = [label]
        total_h = len(lines) * (line_h + 2) - 2
        ty = y0 + max(pad, (y1 - y0 - total_h) // 2)
        for ln in lines:
            draw.text((x0 + pad, ty), ln, font=time_font, fill="black")
            ty += line_h + 2

        for d in range(n_days):
            cx0, cy0, cx1, cy1 = cell_rect(d + 1, si + 1)
            content = grid[si][d]
            max_w = (cx1 - cx0) - 2 * pad
            clines = _wrap_lines_to_width(draw, content, cell_font, max_w)
            lh = draw.textbbox((0, 0), "Ay", font=cell_font)[3] - draw.textbbox((0, 0), "Ay", font=cell_font)[1]
            cy = cy0 + pad
            for ln in clines:
                if cy + lh > cy1 - pad:
                    draw.text((cx0 + pad, cy), "…", font=cell_font, fill="black")
                    break
                draw.text((cx0 + pad, cy), ln, font=cell_font, fill="black")
                cy += lh + 2

    safe = re.sub(r"[^\w\-]+", "_", title, flags=re.UNICODE)[:80]
    out_path = os.path.abspath(f"schedule_{safe}.png")
    img.save(out_path, format="PNG")
    return out_path
